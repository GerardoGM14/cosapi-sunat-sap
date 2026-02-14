import openpyxl
from src.logger.colored_logger import ColoredLogger, Colors
from src.utils.date_current import dateCurrent
from src.socket_client.manager import socket_manager as io
from src.schemas.ISocket import EmitEvent

logger = ColoredLogger()


def filtrar_excel_por_fecha(path_excel: str, fecha_objetivo: str) -> None:
    try:
        nombre_columna: str = "Fecha de emisión de CP"
        
        wb = openpyxl.load_workbook(path_excel)
        ws = wb.active

        col_idx = None
        if isinstance(nombre_columna, int):
            col_idx = nombre_columna
        else:
            for cell in ws[1]:
                if cell.value == nombre_columna:
                    col_idx = cell.column
                    break
        
        if not col_idx:
            logger.log(f"❌ Error: No se encontró la columna '{nombre_columna}'", color=Colors.RED)
            return

        registros_iniciales = ws.max_row - 1
        eliminados = 0
        
        for row_num in range(ws.max_row, 1, -1):
            valor_celda = str(ws.cell(row=row_num, column=col_idx).value)
            
            if fecha_objetivo not in valor_celda:
                ws.delete_rows(row_num)
                eliminados += 1

        wb.save(path_excel)
        
        registros_restantes = registros_iniciales - eliminados

        if registros_restantes > 0:
            logger.log(f"✅ Proceso exitoso.", color=Colors.GREEN)
            logger.log(f"📊 Registros iniciales: {registros_iniciales}", color=Colors.CYAN)
            logger.log(f"🗑️  Registros eliminados: {eliminados}", color=Colors.YELLOW)
            logger.log(f"📋 Registros mantenidos (Fecha {fecha_objetivo}): {registros_restantes}", color=Colors.CYAN)
            
            io.emit(
                event=EmitEvent.SUNAT,
                data={'message': f'Proceso completado. {registros_restantes} registros mantenidos.', 'date': dateCurrent()}
            )
        else:
            logger.log(f"⚠️  Atención: No se encontraron registros para la fecha '{fecha_objetivo}'.", color=Colors.YELLOW)
            logger.log(f"❗ El archivo ahora solo contiene la fila de cabeceras.", color=Colors.RED)
            
            io.emit(
                event=EmitEvent.SUNAT,
                data={'message': f'No se encontraron registros para la fecha {fecha_objetivo}.', 'date': dateCurrent()}
            )
    except Exception as e:
        logger.log(f"❌ Error con filtrado de Excel: {e}", color=Colors.RED)


if __name__ == "__main__":
    filtrar_excel_por_fecha(
        r"D:\cosapi\output\sunat\Pendientes_20100082391_2026_01.xlsx",
        "17/01/2026",
        "Fecha de emisión de CP"
    )
